from openpyxl import Workbook

from model.product import Produto
from model.category import Categoria

class WorkbookFactory:
    path_to_workbook = "data/ControleEstoque.xlsx"
    
    def __init__(self):
        pass

    def _create_generic_sheet(self, book: Workbook, sheet_name: str, model):
        book.create_sheet(sheet_name)
        
        selected_sheet = book[sheet_name]

        # Gera cabeçalho
        atribute_list = list(vars(model).keys())
        selected_sheet.append([*atribute_list])
    
    def _create_sheet_produtos(self, book):
        produto = Produto(None, None)
        self._create_generic_sheet(book, produto.get_table_name(), produto)      
        
    def _create_sheet_categorias(self, book):
        categoria = Categoria(None)
        self._create_generic_sheet(book, categoria.get_table_name(), categoria)
    
    def setup_new_workbook(self):
        book = Workbook()
        book.remove(book["Sheet"])
        self._create_sheet_produtos(book)
        self._create_sheet_categorias(book)
        book.save(self.path_to_workbook)
        