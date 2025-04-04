import os

from openpyxl import Workbook, load_workbook

from service.workbook_factory import WorkbookFactory


class WorkbookLoader:
    def __init__(self):
        pass

    def get_or_create_sheet(file_path, sheet_name):
        """
        Carrega uma planilha existente ou cria uma nova se não existir.
        
        Args:
            file_path (str): Caminho do arquivo Excel
            sheet_name (str): Nome da planilha desejada
        
        Returns:
            tuple: (workbook, worksheet)
        """
        # Verifica se o arquivo existe
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            
            # Verifica se a planilha existe no arquivo
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                print(f"Planilha '{sheet_name}' carregada com sucesso.")
            else:
                ws = wb.create_sheet(sheet_name)
                print(f"Nova planilha '{sheet_name}' criada no arquivo existente.")
        else:
            WorkbookFactory().setup_new_workbook()
        
        return wb, ws